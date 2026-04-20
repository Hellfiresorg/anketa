import csv
import io
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


def _fmt_dt(dt: datetime | None) -> str:
    if not dt:
        return ""
    return dt.strftime("%Y-%m-%d %H:%M")


def _auto_size_columns(ws) -> None:
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                lines = str(cell.value).split('\n')
                cell_max = max(len(line) for line in lines)
                if cell_max > max_len:
                    max_len = cell_max
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 80)


def generate_csv_summary(
    invitations: list[dict],
    questions: list[dict],
    all_responses: list[list[dict]],
) -> tuple[bytes, str, str]:
    buf = io.StringIO()
    writer = csv.writer(buf, delimiter=';')

    headers = ["ID", "Клиент", "Телефон", "Менеджер", "Статус", "Создано", "Заполнено"]
    headers += [f"Вопрос {q['order_index']}" for q in questions]
    writer.writerow(headers)

    for inv, responses in zip(invitations, all_responses):
        resp_map = {r["question_id"]: r["transcription"] for r in responses}
        row = [
            inv["id"],
            inv["client_name"],
            inv.get("client_phone", ""),
            inv.get("manager_full_name", ""),
            inv["status"],
            _fmt_dt(inv.get("created_at")),
            _fmt_dt(inv.get("completed_at")),
        ]
        for q in questions:
            row.append(resp_map.get(q["id"], ""))
        writer.writerow(row)

    content = buf.getvalue().encode("utf-8-sig")
    filename = f"survey_summary_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.csv"
    return content, filename, "text/csv; charset=utf-8"


def generate_csv_rows(
    invitation: dict,
    questions: list[dict],
    responses: list[dict],
) -> tuple[bytes, str, str]:
    buf = io.StringIO()
    writer = csv.writer(buf, delimiter=';')
    writer.writerow([
        "ID приглашения", "Клиент", "Телефон",
        "Номер вопроса", "Текст вопроса", "Транскрипция",
        "Длительность, сек", "Дата записи",
    ])

    resp_map = {r["question_id"]: r for r in responses}
    for q in questions:
        r = resp_map.get(q["id"], {})
        writer.writerow([
            invitation["id"],
            invitation["client_name"],
            invitation.get("client_phone", ""),
            q["order_index"],
            q["text"],
            r.get("transcription", ""),
            r.get("audio_duration_sec", ""),
            _fmt_dt(r.get("created_at")),
        ])

    content = buf.getvalue().encode("utf-8-sig")
    filename = f"survey_rows_{invitation['id']}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.csv"
    return content, filename, "text/csv; charset=utf-8"


def generate_xlsx_summary(
    invitations: list[dict],
    questions: list[dict],
    all_responses: list[list[dict]],
) -> tuple[bytes, str, str]:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Сводка"

    headers = ["ID", "Клиент", "Телефон", "Менеджер", "Статус", "Создано", "Заполнено"]
    headers += [f"Вопрос {q['order_index']}" for q in questions]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    for row_idx, (inv, responses) in enumerate(zip(invitations, all_responses), 2):
        resp_map = {r["question_id"]: r["transcription"] for r in responses}
        row_data = [
            inv["id"], inv["client_name"], inv.get("client_phone", ""),
            inv.get("manager_full_name", ""), inv["status"],
            _fmt_dt(inv.get("created_at")), _fmt_dt(inv.get("completed_at")),
        ]
        for q in questions:
            row_data.append(resp_map.get(q["id"], ""))
        for col, val in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col, value=val)

    ws.row_dimensions[1].height = 25
    for rr in range(2, ws.max_row + 1):
        ws.row_dimensions[rr].height = 40
    _auto_size_columns(ws)

    buf = io.BytesIO()
    wb.save(buf)
    content = buf.getvalue()
    filename = f"survey_summary_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return content, filename, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def generate_xlsx_rows(
    invitation: dict,
    questions: list[dict],
    responses: list[dict],
) -> tuple[bytes, str, str]:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Построчно"

    headers = ["ID приглашения", "Клиент", "Телефон", "Номер вопроса", "Текст вопроса", "Транскрипция", "Длительность, сек", "Дата записи"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    resp_map = {r["question_id"]: r for r in responses}
    for row_idx, q in enumerate(questions, 2):
        r = resp_map.get(q["id"], {})
        row_data = [
            invitation["id"], invitation["client_name"], invitation.get("client_phone", ""),
            q["order_index"] + 1, q["text"],
            r.get("transcription", ""), r.get("audio_duration_sec", ""),
            _fmt_dt(r.get("created_at")),
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.alignment = Alignment(wrap_text=True)

    ws.row_dimensions[1].height = 25
    for rr in range(2, ws.max_row + 1):
        ws.row_dimensions[rr].height = 80
    _auto_size_columns(ws)

    buf = io.BytesIO()
    wb.save(buf)
    content = buf.getvalue()
    filename = f"survey_rows_{invitation['id']}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return content, filename, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def generate_pdf(
    invitation: dict,
    questions: list[dict],
    responses: list[dict],
) -> tuple[bytes, str, str]:
    from weasyprint import HTML

    resp_map = {r["question_id"]: r for r in responses}

    qa_blocks = ""
    for q in questions:
        r = resp_map.get(q["id"], {})
        transcription = r.get("transcription") or "<em>Транскрипция не готова</em>"
        duration = r.get("audio_duration_sec")
        duration_str = f" ({duration:.1f} сек)" if duration else ""
        qa_blocks += f"""
        <div class="qa-block">
            <div class="question">Вопрос {q['order_index']}: {q['text']}</div>
            <div class="answer">{transcription}{duration_str}</div>
        </div>
        """

    html = f"""<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="utf-8">
<style>
  body {{ font-family: Arial, sans-serif; margin: 40px; color: #333; }}
  .header {{ border-bottom: 2px solid #102f6e; padding-bottom: 16px; margin-bottom: 24px; }}
  .header h1 {{ color: #102f6e; margin: 0 0 8px; }}
  .meta {{ font-size: 14px; color: #555; }}
  .qa-block {{ margin: 20px 0; padding: 16px; border-left: 4px solid #102f6e; background: #f9f9f9; }}
  .question {{ font-weight: bold; margin-bottom: 8px; color: #102f6e; }}
  .answer {{ line-height: 1.6; }}
</style>
</head>
<body>
  <div class="header">
    <h1>Анкета клиента</h1>
    <div class="meta">
      <strong>Клиент:</strong> {invitation['client_name']}<br>
      <strong>Телефон:</strong> {invitation.get('client_phone') or '—'}<br>
      <strong>Менеджер:</strong> {invitation.get('manager_full_name', '')}<br>
      <strong>Статус:</strong> {invitation['status']}<br>
      <strong>Создано:</strong> {_fmt_dt(invitation.get('created_at'))}<br>
      <strong>Заполнено:</strong> {_fmt_dt(invitation.get('completed_at'))}
    </div>
  </div>
  {qa_blocks}
</body>
</html>"""

    pdf_bytes = HTML(string=html).write_pdf()
    filename = f"survey_{invitation['id']}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.pdf"
    return pdf_bytes, filename, "application/pdf"
